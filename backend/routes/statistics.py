from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import db, Activity, Registration, CheckIn, User
from sqlalchemy import func
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from utils.auth_helper import parse_user_id

statistics_bp = Blueprint('statistics', __name__)

@statistics_bp.route('/activity/<int:activity_id>', methods=['GET'])
@jwt_required()
def get_activity_statistics(activity_id):
    """获取活动统计（组织者）"""
    user_id = parse_user_id(get_jwt_identity())
    
    if user_id is None:
        return jsonify({'code': 403, 'message': '仅组织者可访问'}), 403
    
    activity = Activity.query.get(activity_id)
    
    if not activity:
        return jsonify({'code': 404, 'message': '活动不存在'}), 404
    
    if activity.organizer_id != user_id:
        return jsonify({'code': 403, 'message': '权限不足'}), 403
    
    # 统计报名人数
    total_registrations = Registration.query.filter_by(
        activity_id=activity_id
    ).filter(Registration.status != 'cancelled').count()
    
    # 统计签到人数
    total_checkins = CheckIn.query.filter_by(activity_id=activity_id).count()
    
    # 计算签到率
    check_in_rate = round(total_checkins / total_registrations * 100, 2) if total_registrations > 0 else 0
    
    # 报名趋势（按天统计）
    registration_trend = db.session.query(
        func.date(Registration.registered_at).label('date'),
        func.count(Registration.id).label('count')
    ).filter_by(activity_id=activity_id).filter(
        Registration.status != 'cancelled'
    ).group_by(func.date(Registration.registered_at)).all()
    
    return jsonify({
        'code': 200,
        'message': '获取成功',
        'data': {
            'activityId': activity_id,
            'activityTitle': activity.title,
            'totalRegistrations': total_registrations,
            'totalCheckIns': total_checkins,
            'checkInRate': check_in_rate,
            'registrationTrend': [
                {'date': str(item.date), 'count': item.count}
                for item in registration_trend
            ]
        }
    })


@statistics_bp.route('/organizer', methods=['GET'])
@jwt_required()
def get_organizer_statistics():
    """获取组织者总体统计"""
    user_id = parse_user_id(get_jwt_identity())
    
    if user_id is None:
        return jsonify({'code': 403, 'message': '仅组织者可访问'}), 403
    
    user = User.query.get(user_id)
    
    if not user or user.role != 'organizer':
        return jsonify({'code': 403, 'message': '权限不足'}), 403
    
    # 统计总活动数
    total_activities = Activity.query.filter_by(organizer_id=user_id).count()
    
    # 统计总报名数
    activity_ids = [a.id for a in Activity.query.filter_by(organizer_id=user_id).all()]
    total_registrations = Registration.query.filter(
        Registration.activity_id.in_(activity_ids),
        Registration.status != 'cancelled'
    ).count() if activity_ids else 0
    
    # 统计总签到数
    total_checkins = CheckIn.query.filter(
        CheckIn.activity_id.in_(activity_ids)
    ).count() if activity_ids else 0
    
    # 计算平均签到率
    average_check_in_rate = round(total_checkins / total_registrations * 100, 2) if total_registrations > 0 else 0
    
    # 最近的活动
    recent_activities = Activity.query.filter_by(organizer_id=user_id).order_by(
        Activity.created_at.desc()
    ).limit(5).all()
    
    return jsonify({
        'code': 200,
        'message': '获取成功',
        'data': {
            'totalActivities': total_activities,
            'totalRegistrations': total_registrations,
            'totalCheckIns': total_checkins,
            'averageCheckInRate': average_check_in_rate,
            'recentActivities': [activity.to_dict() for activity in recent_activities]
        }
    })


@statistics_bp.route('/export/<int:activity_id>', methods=['GET'])
@jwt_required()
def export_statistics(activity_id):
    """导出统计数据为Excel（组织者）"""
    user_id = parse_user_id(get_jwt_identity())
    
    if user_id is None:
        return jsonify({'code': 403, 'message': '仅组织者可访问'}), 403
    
    activity = Activity.query.get(activity_id)
    
    if not activity:
        return jsonify({'code': 404, 'message': '活动不存在'}), 404
    
    if activity.organizer_id != user_id:
        return jsonify({'code': 403, 'message': '权限不足'}), 403
    
    # 判断活动状态
    now = datetime.utcnow()
    if now < activity.start_time:
        activity_status = 'upcoming'
    elif now >= activity.start_time and now <= activity.end_time:
        activity_status = 'ongoing'
    else:
        activity_status = 'completed'
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    
    # 设置标题样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    if activity_status == 'completed':
        # 已结束活动：导出签到数据
        ws.title = '签到统计'
        
        # 设置表头
        headers = ['序号', '姓名', '邮箱', '报名时间', '签到时间', '签到方式', '签到状态']
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 获取所有报名记录（包括已签到和未签到）
        registrations = Registration.query.filter_by(activity_id=activity_id).filter(
            Registration.status != 'cancelled'
        ).order_by(Registration.registered_at).all()
        
        # 填充数据
        for idx, reg in enumerate(registrations, start=1):
            # 查找对应的签到记录
            checkin = CheckIn.query.filter_by(
                activity_id=activity_id,
                user_id=reg.user_id
            ).first()
            
            row = [
                idx,
                reg.user.username,
                reg.user.email,
                reg.registered_at.strftime('%Y-%m-%d %H:%M:%S'),
                checkin.checked_in_at.strftime('%Y-%m-%d %H:%M:%S') if checkin else '',
                '二维码' if checkin and checkin.method == 'qrcode' else ('签到码' if checkin else ''),
                '已签到' if checkin else '未签到'
            ]
            ws.append(row)
        
        # 添加统计信息
        ws.append([])
        total_registrations = len(registrations)
        total_checkins = CheckIn.query.filter_by(activity_id=activity_id).count()
        checkin_rate = round(total_checkins / total_registrations * 100, 2) if total_registrations > 0 else 0
        
        ws.append(['统计信息'])
        ws.append(['总报名人数', total_registrations])
        ws.append(['总签到人数', total_checkins])
        ws.append(['签到率', f'{checkin_rate}%'])
        
    else:
        # 未开始或进行中活动：导出报名数据
        ws.title = '报名统计'
        
        # 设置表头
        headers = ['序号', '姓名', '邮箱', '报名时间', '报名状态', '选择项目']
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 获取所有报名记录
        registrations = Registration.query.filter_by(activity_id=activity_id).filter(
            Registration.status != 'cancelled'
        ).order_by(Registration.registered_at).all()
        
        # 填充数据
        for idx, reg in enumerate(registrations, start=1):
            status_text = '已报名'
            if reg.status == 'checked_in':
                status_text = '已签到'
            
            row = [
                idx,
                reg.user.username,
                reg.user.email,
                reg.registered_at.strftime('%Y-%m-%d %H:%M:%S'),
                status_text,
                reg.sub_item or '无'
            ]
            ws.append(row)
        
        # 添加统计信息
        ws.append([])
        total_registrations = len(registrations)
        ws.append(['统计信息'])
        ws.append(['总报名人数', total_registrations])
        ws.append(['活动状态', '未开始' if activity_status == 'upcoming' else '进行中'])
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    if activity_status == 'completed':
        ws.column_dimensions['G'].width = 12
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f'{activity.title}_{"签到统计" if activity_status == "completed" else "报名统计"}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@statistics_bp.route('/trend', methods=['GET'])
@jwt_required()
def get_registration_trend():
    """获取报名趋势"""
    user_id = parse_user_id(get_jwt_identity())
    
    if user_id is None:
        return jsonify({'code': 403, 'message': '仅组织者可访问'}), 403
    
    user = User.query.get(user_id)
    
    if not user or user.role != 'organizer':
        return jsonify({'code': 403, 'message': '权限不足'}), 403
    
    start_date = request.args.get('startDate')
    end_date = request.args.get('endDate')
    
    if not start_date or not end_date:
        return jsonify({'code': 400, 'message': '缺少日期参数'}), 400
    
    # 获取该组织者的所有活动
    activity_ids = [a.id for a in Activity.query.filter_by(organizer_id=user_id).all()]
    
    if not activity_ids:
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': []
        })
    
    # 统计趋势
    trend = db.session.query(
        func.date(Registration.registered_at).label('date'),
        func.count(Registration.id).label('count')
    ).filter(
        Registration.activity_id.in_(activity_ids),
        Registration.registered_at >= datetime.fromisoformat(start_date),
        Registration.registered_at <= datetime.fromisoformat(end_date),
        Registration.status != 'cancelled'
    ).group_by(func.date(Registration.registered_at)).all()
    
    return jsonify({
        'code': 200,
        'message': '获取成功',
        'data': [
            {'date': str(item.date), 'count': item.count}
            for item in trend
        ]
    })
